"""Parse `SwissHacks Portfolio Construction.xlsx` -> mandates, CIO list, holdings (CLAUDE.md §6).

Sheets used:
  Portfolio Strategies        -> sub-asset-class targets per strategy (Def/Balanced/Growth %, CHF)
  CIO Recommendation List     -> approved universe: BUY/HOLD/SELL + swap candidates
  Sample Portfolio <strategy> -> current vs target positions, incl. Industry Group for same-sector swaps
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from .base import Record


def _num(v: Any) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _opt_num(v: Any) -> Any:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _isodate(v: Any) -> str:
    """Coerce a workbook cell (datetime or string) to an ISO yyyy-mm-dd string."""
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    return str(v)[:10] if v is not None else ""


def _sheet_dicts(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [(h.strip() if isinstance(h, str) else h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        d: dict[str, Any] = {}
        for i, h in enumerate(header):
            if h is None:
                continue
            d[h] = r[i] if i < len(r) else None
        out.append(d)
    return out


class PortfolioWorkbookSource:
    name = "portfolio_xlsx"
    STRATEGIES = ("Defensive", "Balanced", "Growth")

    def __init__(self, path: Path):
        self.path = Path(path)

    def fetch(self, query: Any = None) -> list[Record]:
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        out: list[Record] = []
        out += self._mandates(wb)
        out += self._cio(wb)
        out += self._holdings(wb)
        out += self._transactions(wb)
        out += self._cashflows(wb)
        wb.close()
        return out

    def _mandates(self, wb) -> list[Record]:
        if "Portfolio Strategies" not in wb.sheetnames:
            return []
        recs: list[Record] = []
        pct_col = {"Defensive": "Def %", "Balanced": "Balanced %", "Growth": "Growth %"}
        chf_col = {"Defensive": "Def (CHF)", "Balanced": "Balanced (CHF)", "Growth": "Growth (CHF)"}
        for row in _sheet_dicts(wb["Portfolio Strategies"]):
            ac = row.get("Asset Class")
            sac = row.get("Sub-Asset Class")
            if not ac or not sac:
                continue
            # Skip the footer summary rows ("TOTAL / GLOBAL MANDATE", "Target amount / 10000000"):
            # they are not real sleeves and would otherwise seed a phantom 100pp drift breach.
            if str(ac).strip() in ("TOTAL", "Target amount"):
                continue
            for strat in self.STRATEGIES:
                recs.append(Record(
                    kind="mandate",
                    source_type="mandate",
                    source_id=f"{strat}:{sac}",
                    excerpt=f"{strat} target for {sac}: {row.get(pct_col[strat])}%",
                    payload={
                        "strategy": strat,
                        "asset_class": str(ac),
                        "sub_asset_class": str(sac),
                        "benchmark": str(row.get("Benchmark / Index Reference") or ""),
                        "target_pct": _num(row.get(pct_col[strat])),
                        "target_chf": _num(row.get(chf_col[strat])),
                    },
                ))
        return recs

    def _cio(self, wb) -> list[Record]:
        if "CIO Recommendation List" not in wb.sheetnames:
            return []
        recs: list[Record] = []
        for row in _sheet_dicts(wb["CIO Recommendation List"]):
            isin = row.get("ISIN")
            rating = row.get("Rating")
            if not isin or not rating:
                continue
            recs.append(Record(
                kind="cio",
                source_type="cio_list",
                source_id=str(isin),
                excerpt=f"{rating} · {row.get('Issuer / Asset')} · {row.get('CIO View')}",
                payload={
                    "rating": str(rating).strip().upper(),
                    "asset_class": str(row.get("Asset Class") or ""),
                    "sub_asset_class": str(row.get("Sub-Asset Class") or ""),
                    "region": str(row.get("Region") or ""),
                    "industry_group": str(row.get("Industry Group") or ""),
                    "issuer": str(row.get("Issuer / Asset") or ""),
                    "security": str(row.get("Security / Details") or ""),
                    "isin": str(isin),
                    "cio_view": str(row.get("CIO View") or ""),
                    "valor": str(row.get("Valor") or "") or None,
                    "mic": str(row.get("MIC") or "") or None,
                    "yahoo": str(row.get("Yahoo Ticker") or "") or None,
                },
            ))
        return recs

    def _holdings(self, wb) -> list[Record]:
        recs: list[Record] = []
        for strat in self.STRATEGIES:
            sheet = f"Sample Portfolio {strat}"
            if sheet not in wb.sheetnames:
                continue
            for row in _sheet_dicts(wb[sheet]):
                isin = row.get("ISIN")
                issuer = row.get("Issuer / Asset")
                if not isin or not issuer:
                    continue
                recs.append(Record(
                    kind="holding",
                    source_type="portfolio",
                    source_id=f"{strat}:{isin}",
                    excerpt=f"{issuer} ({isin}) in {strat}",
                    payload={
                        "portfolio": strat,
                        "asset_class": str(row.get("Asset Class") or ""),
                        "sub_asset_class": str(row.get("Sub-Asset Class") or ""),
                        "region": str(row.get("Region") or ""),
                        "industry_group": str(row.get("Industry Group") or ""),
                        "issuer": str(issuer),
                        "security": str(row.get("Security / Details") or ""),
                        "isin": str(isin),
                        "target_chf": _num(row.get("Target (CHF)")),
                        "current_chf": _num(row.get("Current (CHF)")),
                        "valor": str(row.get("Valor") or "") or None,
                        "mic": str(row.get("MIC") or "") or None,
                        "yahoo": str(row.get("Yahoo Ticker") or "") or None,
                    },
                ))
        return recs

    def _transactions(self, wb) -> list[Record]:
        """Historical trades per strategy (Transactions {Defensive,Balanced,Growth}) — cost basis,
        holding period, SELL precedents, and the RM's recorded rationale (HI4)."""
        recs: list[Record] = []
        for strat in self.STRATEGIES:
            sheet = f"Transactions {strat}"
            if sheet not in wb.sheetnames:
                continue
            for row in _sheet_dicts(wb[sheet]):
                txid = row.get("Transaction ID")
                isin = row.get("ISIN")
                side = row.get("Side")
                if not txid or not isin or not side:
                    continue
                issuer = str(row.get("Issuer / Asset") or "")
                rationale = str(row.get("Rationale") or "") or None
                recs.append(Record(
                    kind="transaction", source_type="portfolio", source_id=str(txid),
                    excerpt=(rationale or f"{side} {issuer} ({isin})"),
                    payload={
                        "transaction_id": str(txid),
                        "timestamp": _isodate(row.get("Timestamp")),
                        "portfolio": str(row.get("Portfolio") or strat),
                        "isin": str(isin),
                        "issuer": issuer,
                        "side": str(side).strip().upper(),
                        "quantity": _opt_num(row.get("Quantity")),
                        "price_local": _opt_num(row.get("Price (local)")),
                        "currency": str(row.get("Currency") or "") or None,
                        "fx_chf": _opt_num(row.get("FX → CHF")),
                        "price_chf": _opt_num(row.get("Price (CHF)")),
                        "amount_chf": _num(row.get("Amount (CHF)")),
                        "rationale": rationale,
                        "price_source": str(row.get("Price Source") or "") or None,
                    },
                ))
        return recs

    def _cashflows(self, wb) -> list[Record]:
        """Cash movements (Cash Flows tab): COUPON income, DEPOSIT/WITHDRAWAL flows, FEE drag (HI4)."""
        if "Cash Flows" not in wb.sheetnames:
            return []
        recs: list[Record] = []
        for row in _sheet_dicts(wb["Cash Flows"]):
            fid = row.get("Flow ID")
            side = row.get("Side")
            if not fid or not side:
                continue
            rationale = str(row.get("Rationale") or "") or None
            recs.append(Record(
                kind="cash_flow", source_type="portfolio", source_id=str(fid),
                excerpt=(rationale or f"{side} {row.get('Amount (CHF)')}"),
                payload={
                    "flow_id": str(fid),
                    "timestamp": _isodate(row.get("Timestamp")),
                    "portfolio": str(row.get("Portfolio") or ""),
                    "side": str(side).strip().upper(),
                    "amount_chf": _num(row.get("Amount (CHF)")),
                    "rationale": rationale,
                },
            ))
        return recs
