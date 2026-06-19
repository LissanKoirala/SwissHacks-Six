"""Parse `SwissHacks CRM.xlsx` -> immutable meeting_log records (CLAUDE.md §6).
One tab per client; columns: Date | Medium | RM Name | Client Contact | Note."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from .base import Record


def _date(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    return s[:10]  # YYYY-MM-DD


class CRMWorkbookSource:
    name = "crm_xlsx"

    def __init__(self, path: Path, sheet_to_client: dict[str, str]):
        self.path = Path(path)
        self.sheet_to_client = sheet_to_client

    def fetch(self, query: Any = None) -> list[Record]:
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        out: list[Record] = []
        for sheet, client_id in self.sheet_to_client.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            n = 0
            for r in rows[1:]:
                if not r or r[0] is None:
                    continue
                date = _date(r[0])
                medium = (r[1] or "").strip() if len(r) > 1 else ""
                rm = (r[2] or "").strip() if len(r) > 2 else ""
                contact = (r[3] or "").strip() if len(r) > 3 else ""
                note = (r[4] or "").strip() if len(r) > 4 else ""
                if not note:
                    continue
                n += 1
                out.append(Record(
                    kind="meeting_log",
                    source_type="crm_log",
                    source_id=f"{client_id}#{date}#{n}",
                    excerpt=note[:240],
                    payload={
                        "client_id": client_id,
                        "timestamp": date,
                        "modality": medium,
                        "rm_name": rm,
                        "contact": contact,
                        "note": note,
                        "sheet": sheet,
                    },
                ))
        wb.close()
        return out
