#!/usr/bin/env python3
"""Build an Obsidian-style node graph (single self-contained HTML) from the SwissHacks CRM workbook.

Reads ../data/SwissHacks CRM.xlsx, derives a graph of:
  RM -> families -> people -> interactions -> {medium, theme}
and writes crm-graph.html with the data baked in (works offline, no CDN).
"""
import json
import re
from pathlib import Path
from datetime import datetime, date

import openpyxl

HERE = Path(__file__).resolve().parent
CRM_XLSX = HERE.parent / "data" / "SwissHacks CRM.xlsx"
OUT_HTML = HERE / "crm-graph.html"

# Map workbook sheet -> family display name.
FAMILIES = {
    "CRM Raeber": "Räber",
    "CRM Schneider": "Schneider",
    "CRM Huber": "Huber",
    "CRM Ammann": "Ammann",
}

# Curated themes: title -> list of lowercase substrings to match in the note text.
THEMES = {
    "Capital Preservation": ["preservation", "preserv", "conservative", "sleep at night", "steward", "defensive"],
    "Dividends / Income": ["dividend", "payout", "cash flow", "cashflow", "income", "yield"],
    "ESG / Sustainability": ["esg", "sustainab", "biodiversity", "reforest", "nature", "greenwash", "ecosystem", "ngo", "philanthrop"],
    "Supply-Chain Governance": ["supply-chain", "supply chain", "sweatshop", "labor", "labour", "wage theft", "governance liabilit", "exploitation"],
    "Succession Planning": ["succession", "grandchildren", "wealth transmission", "transmission", "children"],
    "Liquidity Event": ["withdrawal", "capital call", "deposit", "renovation", "acquisition", "top-up"],
    "Anti-Speculation": ["speculat", "high-beta", "bubble", "tail risk", "asymmetric"],
    "Reputation Risk": ["reputation", "hypocrisy", "backlash", "public face", "brand equity", "name is linked"],
}

# Stable colour per node type (matches the legend in the HTML).
TYPE_COLOR = {
    "rm": "#e0b3ff",
    "family": "#ffd166",
    "person": "#4cc9f0",
    "medium": "#76c893",
    "interaction": "#9aa0b5",
    "theme": "#f08080",
}


def norm_people(contact: str):
    """Split a Client Contact cell into individual person names."""
    contact = (contact or "").strip()
    if not contact:
        return []
    # Handle "Hubertus & Carmen", "Marius & Elena", etc. -> attach family surname later.
    parts = re.split(r"\s*&\s*|\s+and\s+", contact)
    return [p.strip() for p in parts if p.strip()]


def main():
    wb = openpyxl.load_workbook(CRM_XLSX, data_only=True)

    nodes = {}   # id -> node dict
    links = []   # {source, target}
    link_seen = set()

    def add_node(node_id, **attrs):
        if node_id not in nodes:
            nodes[node_id] = {"id": node_id, **attrs}
        return node_id

    def add_link(a, b):
        key = (a, b) if a < b else (b, a)
        if key not in link_seen and a != b:
            link_seen.add(key)
            links.append({"source": a, "target": b})

    rm_id = add_node("rm:Thomas Keller", label="Thomas Keller", type="rm",
                     detail="Relationship Manager for all client families.")

    for sheet, family in FAMILIES.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        fam_id = add_node(f"family:{family}", label=f"{family} Family", type="family",
                          detail=f"Client household: {family}.")
        add_link(rm_id, fam_id)

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for idx, row in enumerate(rows, start=1):
            if not row or row[0] is None:
                continue
            d, medium, rm_name, contact, note = (list(row) + [None] * 5)[:5]

            # Date label.
            if isinstance(d, (datetime, date)):
                date_label = d.strftime("%Y-%m-%d")
            else:
                date_label = str(d) if d else ""

            medium = (medium or "Other").strip()
            note = (note or "").strip()

            # People involved.
            people = norm_people(contact)
            person_ids = []
            for p in people:
                # Give surname-less first names a family suffix for uniqueness/clarity.
                disp = p if family.split()[0] in p or len(p.split()) > 1 else f"{p} {family}"
                pid = add_node(f"person:{disp}", label=disp, type="person",
                               detail=f"{family} household member.")
                add_link(fam_id, pid)
                person_ids.append(pid)

            # Medium node (shared across all families).
            mid = add_node(f"medium:{medium}", label=medium, type="medium",
                           detail=f"Contact channel: {medium}.")

            # Interaction node.
            iid = f"int:{sheet}:{idx}"
            short = note[:60].rsplit(" ", 1)[0] + "…" if len(note) > 60 else note
            add_node(iid, label=f"{date_label} · {medium}", type="interaction",
                     date=date_label, medium=medium, family=family,
                     contact=contact or "", detail=note, summary=short)
            add_link(fam_id, iid)
            add_link(mid, iid)
            for pid in person_ids:
                add_link(pid, iid)

            # Theme tagging.
            low = note.lower()
            for theme, kws in THEMES.items():
                if any(kw in low for kw in kws):
                    tid = add_node(f"theme:{theme}", label=theme, type="theme",
                                   detail=f"Cross-client theme: {theme}.")
                    add_link(iid, tid)

    # Apply colours + degree (for node sizing).
    deg = {}
    for l in links:
        deg[l["source"]] = deg.get(l["source"], 0) + 1
        deg[l["target"]] = deg.get(l["target"], 0) + 1
    node_list = []
    for n in nodes.values():
        n["color"] = TYPE_COLOR.get(n["type"], "#888")
        n["degree"] = deg.get(n["id"], 0)
        node_list.append(n)

    graph = {"nodes": node_list, "links": links}
    html = HTML_TEMPLATE.replace("__GRAPH_DATA__", json.dumps(graph, ensure_ascii=False))
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"Wrote {OUT_HTML}")
    print(f"  {len(node_list)} nodes, {len(links)} links")
    from collections import Counter
    c = Counter(n["type"] for n in node_list)
    print("  by type:", dict(c))


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>CRM Graph — Obsidian style</title>
<style>
  :root { --bg:#1a1b26; --panel:#24283b; --text:#c0caf5; --muted:#565f89; --accent:#7aa2f7; }
  * { box-sizing: border-box; }
  html,body { margin:0; height:100%; background:var(--bg); color:var(--text);
              font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; overflow:hidden; }
  #canvas { display:block; cursor:grab; }
  #canvas:active { cursor:grabbing; }
  .panel { position:fixed; background:rgba(36,40,59,.92); backdrop-filter:blur(8px);
           border:1px solid #2f344d; border-radius:12px; padding:14px 16px; box-shadow:0 8px 30px rgba(0,0,0,.4); }
  #controls { top:16px; left:16px; width:260px; }
  #controls h1 { margin:0 0 4px; font-size:15px; letter-spacing:.3px; }
  #controls .sub { color:var(--muted); font-size:11px; margin-bottom:10px; }
  #search { width:100%; padding:8px 10px; border-radius:8px; border:1px solid #343a57;
            background:#1f2335; color:var(--text); font-size:13px; outline:none; }
  #search:focus { border-color:var(--accent); }
  .legend { margin-top:12px; display:flex; flex-direction:column; gap:6px; }
  .legend label { display:flex; align-items:center; gap:8px; font-size:12px; cursor:pointer; user-select:none; }
  .legend .dot { width:11px; height:11px; border-radius:50%; flex:none; }
  .legend input { accent-color:var(--accent); }
  .legend .count { color:var(--muted); margin-left:auto; font-size:11px; }
  #detail { bottom:16px; left:16px; width:340px; max-height:42vh; overflow:auto; display:none; }
  #detail .tag { display:inline-block; font-size:10px; padding:2px 8px; border-radius:999px;
                 background:#1f2335; color:var(--muted); margin-bottom:8px; text-transform:uppercase; letter-spacing:.6px; }
  #detail h2 { margin:0 0 6px; font-size:14px; }
  #detail .meta { color:var(--muted); font-size:12px; margin-bottom:8px; }
  #detail p { margin:0; font-size:13px; line-height:1.5; }
  #hint { bottom:16px; right:16px; color:var(--muted); font-size:11px; text-align:right; line-height:1.6; }
  #tooltip { position:fixed; pointer-events:none; background:#0f1018; border:1px solid #343a57;
             padding:5px 9px; border-radius:7px; font-size:12px; display:none; max-width:280px; z-index:10; }
  .btn { margin-top:10px; width:100%; padding:7px; border-radius:8px; border:1px solid #343a57;
         background:#1f2335; color:var(--text); font-size:12px; cursor:pointer; }
  .btn:hover { border-color:var(--accent); }
</style>
</head>
<body>
<canvas id="canvas"></canvas>

<div id="controls" class="panel">
  <h1>CRM Knowledge Graph</h1>
  <div class="sub">SwissHacks · client relationship web</div>
  <input id="search" placeholder="Search nodes…" autocomplete="off"/>
  <div class="legend" id="legend"></div>
  <button class="btn" id="reset">Reset view</button>
</div>

<div id="detail" class="panel">
  <span class="tag" id="d-type"></span>
  <h2 id="d-title"></h2>
  <div class="meta" id="d-meta"></div>
  <p id="d-body"></p>
</div>

<div id="hint">drag node · scroll zoom · drag bg pan · click node</div>
<div id="tooltip"></div>

<script>
const GRAPH = __GRAPH_DATA__;
const TYPE_LABELS = { rm:"Relationship Mgr", family:"Family", person:"Person", medium:"Medium", interaction:"Interaction", theme:"Theme" };
const TYPE_COLOR = { rm:"#e0b3ff", family:"#ffd166", person:"#4cc9f0", medium:"#76c893", interaction:"#9aa0b5", theme:"#f08080" };

const canvas = document.getElementById('canvas');
const ctx = canvas.getContext('2d');
let W, H, DPR;
function resize() {
  DPR = window.devicePixelRatio || 1;
  W = window.innerWidth; H = window.innerHeight;
  canvas.width = W*DPR; canvas.height = H*DPR;
  canvas.style.width = W+'px'; canvas.style.height = H+'px';
  ctx.setTransform(DPR,0,0,DPR,0,0);
}
window.addEventListener('resize', resize); resize();

// ---- build node/link objects ----
const nodes = GRAPH.nodes.map(n => ({...n,
  x: W/2 + (Math.random()-.5)*Math.min(W,H)*0.8,
  y: H/2 + (Math.random()-.5)*Math.min(W,H)*0.8,
  vx:0, vy:0,
  r: 4 + Math.min(14, Math.sqrt(n.degree||1)*2.4)
}));
const byId = new Map(nodes.map(n => [n.id, n]));
const links = GRAPH.links.map(l => ({source: byId.get(l.source), target: byId.get(l.target)})).filter(l=>l.source&&l.target);

// adjacency for highlight
const adj = new Map(nodes.map(n=>[n.id, new Set()]));
links.forEach(l=>{ adj.get(l.source.id).add(l.target.id); adj.get(l.target.id).add(l.source.id); });

// ---- view transform ----
let view = {x:0, y:0, k:1};
function toScreen(p){ return {x: p.x*view.k + view.x, y: p.y*view.k + view.y}; }
function toWorld(sx,sy){ return {x:(sx-view.x)/view.k, y:(sy-view.y)/view.k}; }

// ---- filters ----
const enabled = new Set(Object.keys(TYPE_LABELS));
const legend = document.getElementById('legend');
const counts = {};
nodes.forEach(n=>counts[n.type]=(counts[n.type]||0)+1);
Object.keys(TYPE_LABELS).forEach(t=>{
  if(!counts[t]) return;
  const lab = document.createElement('label');
  lab.innerHTML = `<input type="checkbox" checked data-t="${t}"><span class="dot" style="background:${TYPE_COLOR[t]}"></span>${TYPE_LABELS[t]}<span class="count">${counts[t]}</span>`;
  legend.appendChild(lab);
});
legend.addEventListener('change', e=>{
  const t = e.target.dataset.t;
  if(!t) return;
  e.target.checked ? enabled.add(t) : enabled.delete(t);
});
const visible = n => enabled.has(n.type);

// ---- force simulation ----
const REPULSE = 5200, SPRING = 0.012, LINK_LEN = 70, CENTER = 0.012, DAMP = 0.86;
let alpha = 1;
function tick(){
  if(alpha < 0.005) { alpha = 0.005; }
  // repulsion (O(n^2); fine for this size)
  for(let i=0;i<nodes.length;i++){
    const a=nodes[i]; if(!visible(a)) continue;
    for(let j=i+1;j<nodes.length;j++){
      const b=nodes[j]; if(!visible(b)) continue;
      let dx=a.x-b.x, dy=a.y-b.y, d2=dx*dx+dy*dy||0.01;
      if(d2>90000) continue;
      const f = REPULSE/d2, d=Math.sqrt(d2);
      const fx=dx/d*f, fy=dy/d*f;
      a.vx+=fx*alpha; a.vy+=fy*alpha; b.vx-=fx*alpha; b.vy-=fy*alpha;
    }
  }
  // springs
  for(const l of links){
    const a=l.source,b=l.target; if(!visible(a)||!visible(b)) continue;
    let dx=b.x-a.x, dy=b.y-a.y, d=Math.sqrt(dx*dx+dy*dy)||0.01;
    const f=(d-LINK_LEN)*SPRING;
    const fx=dx/d*f, fy=dy/d*f;
    a.vx+=fx*alpha; a.vy+=fy*alpha; b.vx-=fx*alpha; b.vy-=fy*alpha;
  }
  // centering + integrate
  for(const n of nodes){
    if(!visible(n)) continue;
    n.vx += (W/2 - n.x)*CENTER*alpha;
    n.vy += (H/2 - n.y)*CENTER*alpha;
    if(n===dragNode) continue;
    n.vx*=DAMP; n.vy*=DAMP;
    n.x+=n.vx; n.y+=n.vy;
  }
  alpha *= 0.992;
}

// ---- interaction state ----
let hoverNode=null, selNode=null, dragNode=null;
let panning=false, last={x:0,y:0}, dragMoved=false;

function nodeAt(sx,sy){
  const w = toWorld(sx,sy);
  let best=null, bd=Infinity;
  for(const n of nodes){
    if(!visible(n)) continue;
    const dx=n.x-w.x, dy=n.y-w.y, d=Math.sqrt(dx*dx+dy*dy);
    if(d < n.r+6/view.k && d<bd){ bd=d; best=n; }
  }
  return best;
}

canvas.addEventListener('mousedown', e=>{
  const n = nodeAt(e.clientX, e.clientY);
  dragMoved=false;
  if(n){ dragNode=n; n.fixed=true; }
  else { panning=true; }
  last={x:e.clientX, y:e.clientY};
});
window.addEventListener('mousemove', e=>{
  const tooltip=document.getElementById('tooltip');
  if(dragNode){
    const w=toWorld(e.clientX,e.clientY);
    dragNode.x=w.x; dragNode.y=w.y; dragNode.vx=0; dragNode.vy=0;
    alpha=Math.max(alpha,0.3); dragMoved=true;
  } else if(panning){
    view.x += e.clientX-last.x; view.y += e.clientY-last.y;
    last={x:e.clientX,y:e.clientY}; dragMoved=true;
  } else {
    const n=nodeAt(e.clientX,e.clientY);
    hoverNode=n;
    if(n){ tooltip.style.display='block'; tooltip.style.left=(e.clientX+12)+'px';
           tooltip.style.top=(e.clientY+12)+'px';
           tooltip.textContent = n.label + (n.type==='interaction' && n.summary ? ' — '+n.summary : ''); }
    else tooltip.style.display='none';
    canvas.style.cursor = n? 'pointer':'grab';
  }
});
window.addEventListener('mouseup', e=>{
  if(dragNode){ dragNode.fixed=false; if(!dragMoved) select(dragNode); dragNode=null; }
  else if(panning && !dragMoved){ const n=nodeAt(e.clientX,e.clientY); n?select(n):select(null); }
  panning=false;
});
canvas.addEventListener('wheel', e=>{
  e.preventDefault();
  const w=toWorld(e.clientX,e.clientY);
  const f = e.deltaY<0 ? 1.12 : 1/1.12;
  view.k = Math.max(0.15, Math.min(5, view.k*f));
  view.x = e.clientX - w.x*view.k;
  view.y = e.clientY - w.y*view.k;
}, {passive:false});

// ---- selection / detail panel ----
const detail=document.getElementById('detail');
function select(n){
  selNode=n;
  if(!n){ detail.style.display='none'; return; }
  detail.style.display='block';
  document.getElementById('d-type').textContent = TYPE_LABELS[n.type]||n.type;
  document.getElementById('d-title').textContent = n.label;
  let meta='';
  if(n.type==='interaction'){
    meta = [n.date, n.medium, n.contact, n.family+' family'].filter(Boolean).join(' · ');
  } else {
    meta = (adj.get(n.id)?.size||0) + ' connections';
  }
  document.getElementById('d-meta').textContent = meta;
  document.getElementById('d-body').textContent = n.detail || '';
}

// ---- search ----
let query='';
document.getElementById('search').addEventListener('input', e=>{ query=e.target.value.trim().toLowerCase(); });

document.getElementById('reset').addEventListener('click', ()=>{
  view={x:0,y:0,k:1}; alpha=1; select(null);
});

// ---- render loop ----
function render(){
  tick();
  ctx.clearRect(0,0,W,H);
  ctx.save();

  const focus = selNode || hoverNode;
  const neigh = focus ? adj.get(focus.id) : null;

  // links
  ctx.lineWidth = 1;
  for(const l of links){
    const a=l.source,b=l.target;
    if(!visible(a)||!visible(b)) continue;
    const pa=toScreen(a), pb=toScreen(b);
    let active = focus && (a===focus||b===focus);
    ctx.strokeStyle = active ? 'rgba(122,162,247,.55)' : 'rgba(120,130,170,.13)';
    ctx.beginPath(); ctx.moveTo(pa.x,pa.y); ctx.lineTo(pb.x,pb.y); ctx.stroke();
  }

  // nodes
  for(const n of nodes){
    if(!visible(n)) continue;
    const p=toScreen(n);
    const r=n.r*Math.sqrt(view.k);
    let dim = false, hit = true;
    if(focus){ dim = !(n===focus || (neigh&&neigh.has(n.id))); }
    if(query){ hit = n.label.toLowerCase().includes(query) || (n.detail||'').toLowerCase().includes(query); }

    ctx.globalAlpha = dim ? 0.18 : (query && !hit ? 0.12 : 1);
    ctx.beginPath(); ctx.arc(p.x,p.y,r,0,Math.PI*2);
    ctx.fillStyle = n.color; ctx.fill();
    if(query && hit){ ctx.lineWidth=2; ctx.strokeStyle='#fff'; ctx.stroke(); }
    if(n===selNode){ ctx.lineWidth=2.5; ctx.strokeStyle='#fff'; ctx.stroke(); }

    // labels: show for big nodes, on focus, or on zoom-in
    const showLabel = (n.r>7 || view.k>1.4 || n===focus || (neigh&&neigh.has(n.id))) && !dim;
    if(showLabel){
      ctx.globalAlpha = dim?0.2:0.92;
      ctx.fillStyle = '#c0caf5';
      ctx.font = `${Math.max(10, Math.min(14, 10*view.k))}px -apple-system, sans-serif`;
      ctx.textAlign='center';
      const txt = n.type==='interaction' ? n.label : n.label;
      ctx.fillText(txt, p.x, p.y - r - 4);
    }
  }
  ctx.globalAlpha=1;
  ctx.restore();
  requestAnimationFrame(render);
}
render();

// gentle reheat on load so it settles nicely
setTimeout(()=>{ alpha=1; }, 50);
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
