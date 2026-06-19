#!/usr/bin/env python3
"""Export the SwissHacks workbooks to JSON for the demo backend's CRM agent layer.

Writes (committed, consumed by demo/src/backend/crm-graph/crm.service.ts):
  demo/src/backend/crm-graph/data/crm.json         clients + their CRM interactions
  demo/src/backend/crm-graph/data/portfolios.json  holdings per sample portfolio strategy
  demo/src/backend/crm-graph/data/cio.json         CIO BUY/HOLD/SELL conviction list
"""
import json
from pathlib import Path
from datetime import datetime, date

import openpyxl

HERE = Path(__file__).resolve().parent
DATA = HERE.parent / "data"
OUT = HERE.parent / "demo" / "src" / "backend" / "crm-graph" / "data"
OUT.mkdir(parents=True, exist_ok=True)

CRM_XLSX = DATA / "SwissHacks CRM.xlsx"
PORT_XLSX = DATA / "SwissHacks Portfolio Construction.xlsx"

# Sheet -> (client id, household display name).
CLIENTS = {
    "CRM Raeber": ("raeber", "Räber"),
    "CRM Schneider": ("schneider", "Schneider"),
    "CRM Huber": ("huber", "Huber"),
    "CRM Ammann": ("ammann", "Ammann"),
}

# Inferred mandate per client, derived from their stated goals in the notes.
# Overridable at query time; this is just the default assignment.
MANDATE = {
    "raeber": "Defensive",
    "schneider": "Balanced",
    "huber": "Balanced",
    "ammann": "Balanced",
}


def iso(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v) if v is not None else None


def export_crm():
    wb = openpyxl.load_workbook(CRM_XLSX, data_only=True)
    clients = []
    for sheet, (cid, name) in CLIENTS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        interactions = []
        contacts = set()
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if not row or row[0] is None:
                continue
            d, medium, rm, contact, note = (list(row) + [None] * 5)[:5]
            if contact:
                contacts.add(str(contact).strip())
            interactions.append({
                "id": f"{cid}-{i:02d}",
                "date": iso(d),
                "medium": (medium or "").strip(),
                "rm": (rm or "").strip(),
                "contact": (contact or "").strip(),
                "note": (note or "").strip(),
            })
        interactions.sort(key=lambda x: x["date"] or "")
        clients.append({
            "id": cid,
            "name": name,
            "household": f"{name} Family",
            "mandate": MANDATE.get(cid),
            "contacts": sorted(contacts),
            "interactionCount": len(interactions),
            "firstContact": interactions[0]["date"] if interactions else None,
            "lastContact": interactions[-1]["date"] if interactions else None,
            "interactions": interactions,
        })
    (OUT / "crm.json").write_text(json.dumps(clients, ensure_ascii=False, indent=2), "utf-8")
    print(f"crm.json: {len(clients)} clients, "
          f"{sum(c['interactionCount'] for c in clients)} interactions")


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def export_portfolios():
    wb = openpyxl.load_workbook(PORT_XLSX, data_only=True)
    sheets = {
        "Defensive": "Sample Portfolio Defensive",
        "Balanced": "Sample Portfolio Balanced",
        "Growth": "Sample Portfolio Growth",
    }
    portfolios = {}
    for strategy, sheet in sheets.items():
        ws = wb[sheet]
        holdings = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            (asset, sub, region, industry, issuer, sec, isin,
             target, current, valor, mic, ticker) = (list(row) + [None] * 12)[:12]
            if not issuer:
                continue
            holdings.append({
                "assetClass": (asset or "").strip(),
                "subAssetClass": (sub or "").strip(),
                "region": (region or "").strip(),
                "industry": (industry or "").strip(),
                "issuer": (issuer or "").strip(),
                "security": (sec or "").strip(),
                "isin": (isin or "").strip(),
                "targetCHF": num(target),
                "currentCHF": num(current),
                "ticker": (ticker or "").strip(),
            })
        portfolios[strategy] = holdings
    (OUT / "portfolios.json").write_text(json.dumps(portfolios, ensure_ascii=False, indent=2), "utf-8")
    print("portfolios.json: " + ", ".join(f"{k}={len(v)}" for k, v in portfolios.items()))


def export_cio():
    wb = openpyxl.load_workbook(PORT_XLSX, data_only=True)
    ws = wb["CIO Recommendation List"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        (rating, since, asset, sub, region, industry, issuer, sec,
         isin, view, valor, mic, ticker, asof) = (list(row) + [None] * 14)[:14]
        if not issuer:
            continue
        rows.append({
            "rating": (rating or "").strip(),
            "assetClass": (asset or "").strip(),
            "industry": (industry or "").strip(),
            "issuer": (issuer or "").strip(),
            "isin": (isin or "").strip(),
            "view": (view or "").strip(),
            "ticker": (ticker or "").strip(),
        })
    (OUT / "cio.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2), "utf-8")
    print(f"cio.json: {len(rows)} rated instruments")


if __name__ == "__main__":
    export_crm()
    export_portfolios()
    export_cio()
    print(f"-> {OUT}")
